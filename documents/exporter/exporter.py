from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import render

import datetime
import json
import uuid
from django.contrib import messages
from datetime import timedelta


class Exporter:

    def ExcelExporter(self, title, headers, body):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"

        # Define border style
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Define header font and fill
        header_font = Font(bold=True, name="Times New Roman")
        cell_font = Font(bold=False, name="Times New Roman")
        header_fill = PatternFill("solid", fgColor="D7E4BC")
        
        # Define alignment and text wrap
        text_alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="centerContinuous")

        # Write the custom title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "T.K. Group"
        title_cell.font = header_font
        title_cell.alignment = header_alignment

        
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        title_cell = ws.cell(row=2, column=1)
        title_cell.value = title
        title_cell.font = header_font
        title_cell.alignment = header_alignment


        # Write headers with formatting
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.alignment = text_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        # Write data with border formatting
        for row_num, data_row in enumerate(body, start=4):
            for col_num, data in enumerate(data_row, start=1):
                if isinstance(data, (datetime.datetime, uuid.UUID)):
                    data = str(data)
                if isinstance(data, (dict, list)):
                    data = json.dumps(data)
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = data
                cell.font = cell_font
                cell.alignment = text_alignment
                cell.border = thin_border

        # Auto fit columns
        for col_num in range(1, len(headers) + 1):
            max_length = max(len(str(ws.cell(row=row, column=col_num).value)) for row in range(1, len(body) + 3))
            ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

        # Save the workbook to the output
        wb.save(output)
        output.seek(0)
        return output

