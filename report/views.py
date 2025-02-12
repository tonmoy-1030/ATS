from django.shortcuts import render, redirect
from django.urls import reverse
from datetime import timedelta
import os
from django.http import HttpResponse,JsonResponse
from .forms import scheduleForm, DailyJoiningForm
import tempfile  # To create temporary files
from pyreportjasper import PyReportJasper
from employees.forms import EmployeeFilter
from employees.models import Employee
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from .models import DailyJoining
from django.views.generic.list import ListView
import pandas as pd
import openpyxl
from django.contrib.auth.decorators import user_passes_test
from .forms import EmployeeReportForm
from django.db.models import Q




os.environ['JAVA_HOME'] = r"C:\Program Files\Java\jdk-17"
os.environ['PATH'] = os.environ['PATH'] + os.pathsep + os.path.join(os.environ['JAVA_HOME'], 'bin')

# Create your views here.


# candidates_token = os.path.join(project_root, 'utls/candidates_token.json')
# credentials_files = os.path.join(project_root, 'utls/credentials.json')

# Define the resources directory
RESOURCES_DIR = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'resources')

# Daily Joining
class DailyJoiningCreateView(CreateView):
    model = DailyJoining
    form_class = DailyJoiningForm
    template_name = 'report/daily_joining_form.html'
    success_url = '/report/daily_joining/'

    def get_form_kwargs(self):
        # Get the default form kwargs and add the user
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Assign the logged-in user to the form instance (if applicable)
        form.instance.user = self.request.user
        return super().form_valid(form)

    
class DailyJoiningUpdateView(UpdateView):
    model = DailyJoining
    form_class = DailyJoiningForm
    template_name = 'report/daily_joining_form.html'
    success_url = '/report/daily_joining/'

    def get_form_kwargs(self):
        # Get the default form kwargs and add the user
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Assign the logged-in user to the form instance (if applicable)
        form.instance.user = self.request.user
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('report:daily_joining_list')
    
class DailyJoiningDeleteView(DeleteView):
    model = DailyJoining
    template_name = 'report/daily_joining_delete.html'
    success_url = '/report/daily_joining/list'
    
class DailyJoiningListView(ListView):
    model = DailyJoining
    template_name = 'report/daily_joining_list.html'
    context_object_name = 'joinings'
    paginate_by = 10
    ordering = ['-date', 'unit']
    
    def get_queryset(self):
        return DailyJoining.objects.filter(user=self.request.user).order_by('-date', 'unit')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Daily Joining List'
        return context

class DailyJoiningReportView(ListView):
    model = DailyJoining
    template_name = 'report/daily_joining_report.html'
    context_object_name = 'joinings'
    paginate_by = 10
    ordering = ['-date', 'unit']
    
    def get_queryset(self):
        print(DailyJoining.objects.all().order_by('-date', 'unit'))
        return DailyJoining.objects.all().order_by('-date', 'unit')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Daily Joining Report'
        return context
import django.utils.timezone as timezone
from django.db.models import Sum    

def get_daily_joinings(request):

    # Get the start_date from the request (default to today's date)
    start_date = request.GET.get('start_date', timezone.now().date()+timedelta(days=1))
    start_date = pd.to_datetime(start_date).date()
    # Calculate the range for the last 10 days

    
    end_date = start_date - timedelta(days=10)
    # Query the data within the 10-day range
    daily_joinings = DailyJoining.objects.filter(
        date__lt=start_date, date__gte=end_date
    ).values(
        'date',
        'unit__short_name',
        'location',
        'employee_category',
        'recruitment_type',
        'joinings_count'
    ).order_by('-date')  # Order by date in descending order

    pd.set_option('future.no_silent_downcasting', True)

    # Convert the queryset to a pandas DataFrame
    df = pd.DataFrame(daily_joinings)
    if not df.empty:
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df['date'] = df['date'].dt.strftime('%d-%b-%y')

        # Create the pivot table with a multi-level index (date, unit, location, category)
        pivot_table = pd.pivot_table(
            df,
            values='joinings_count',
            index=['unit__short_name', 'location', 'employee_category'],
            columns=['date'],
            aggfunc='sum',
            margins=True,
            margins_name='Total',
            fill_value="-"
        )

        # Reset the index to flatten the pivot table (for easier rendering in the template)
        pivot_table = pivot_table.reset_index()

        # Prepare the data for template rendering
        table_data = pivot_table.to_dict('records')
        timestamp_keys = list(table_data[0].keys())[3:-1] if table_data else []
    else:
        table_data = []
        timestamp_keys = []

    # Calculate the next and previous date ranges
    next_start_date = (start_date + timedelta(days=10)).strftime('%Y-%m-%d')
    prev_start_date = (start_date - timedelta(days=10)).strftime('%Y-%m-%d')
    
    
    grouped_data = DailyJoining.objects.filter(date__year=timezone.now().year).values('unit__short_name').annotate(
        total_joinings=Sum('joinings_count')  # Sum the joinings_count for each group
    )  # Order by total_joinings in descending order
    
    grand_total = DailyJoining.objects.filter(date__year=timezone.now().year).aggregate(
        total_joinings=Sum('joinings_count')
    )['total_joinings']
    
    # Render the result in the template
    return render(request, 'report/daily_joining_report.html', {
        'rows': table_data,
        'timestamp_keys': timestamp_keys,
        'current_start_date': (start_date-timedelta(days=1)).strftime('%Y-%m-%d'),
        'next_start_date': next_start_date,
        'prev_start_date': prev_start_date,
        'grouped_data': grouped_data,
        'grand_total': grand_total
    })

def daily_joinings_chart(request):
    
    grouped_data = DailyJoining.objects.filter(date__year=timezone.now().year).values('unit__short_name').annotate(
        total_joinings=Sum('joinings_count')  # Sum the joinings_count for each group
    )  # Order by total_joinings in descending order
    
    labels = [item['unit__short_name'] for item in grouped_data]
    data = [item['total_joinings'] for item in grouped_data]
    return JsonResponse({'labels': labels, 'data': data})


def interviewSchedule(request):
    settings_dir = os.path.dirname(__file__)
    project_root = os.path.abspath(os.path.dirname(settings_dir))
    input_file = os.path.join(project_root, 'report/resources/Interview_schedule.jrxml')

    if request.method == "POST":
        form = scheduleForm(request.POST)
        if form.is_valid():
            from_date = form.cleaned_data.get("from_date")
            to_date = form.cleaned_data.get("to_date")
            unit = form.cleaned_data.get('unit')
            format_ = form.cleaned_data.get("format_")  # The selected format (e.g., 'pdf', 'docx', 'xls')

            # Handle single or multiple units
            if len(unit) > 1:
                unit = tuple(unit)
            else:
                unit = unit[0]  # Take the single unit directly

            # Ensure from_date and to_date are provided
            if not from_date or not to_date:
                return render(request, "report/interviewSchedule.html", {
                    'form': form,
                    'error': 'Please provide both from and to dates.'
                })

            # Format the dates for the report
            from_date_str = from_date.strftime('%Y-%m-%d')
            to_date_str = (to_date + timedelta(days=1)).strftime('%Y-%m-%d')

            # Create PyReportJasper object
            jasper = PyReportJasper()

            # Define parameters for the report
            param = {
                'startDate': from_date_str,
                'endDate': to_date_str,
                'businessUnit': unit
            }

            # Create a temporary file to hold the output
            with tempfile.NamedTemporaryFile(suffix=f'.{format_}', delete=False) as temp_file:
                output_file_path = temp_file.name  # Get the temporary file path

            # Configure the report with MySQL database connection
            jasper.config(
                input_file=input_file,
                output_file=output_file_path,  # Save output to the temporary file
                parameters=param,
                output_formats=[format_],  # Output in the selected format
                db_connection={
                    'driver': 'mysql',
                    'username': 'root',
                    'password': 'Tonmoy1030',
                    'host': 'localhost',
                    'database': 'atsdb',
                    'port': '3306',
                    'jdbc_driver': 'com.mysql.cj.jdbc.Driver',
                    'jdbc_url': 'jdbc:mysql://localhost:3306/atsdb'
                }
            )

            # Process the report and generate the output in the temporary file
            jasper.process_report()

            # Read the generated file from the temporary file
            with open(output_file_path, 'rb') as output_file:
                response = HttpResponse(output_file.read(), content_type='application/octet-stream')
                response['Content-Disposition'] = f'attachment; filename="Interview_Schedule.{format_}"'
                return response
    else:
        form = scheduleForm()

    return render(request, "report/interviewSchedule.html", context={'form': form})

def employee_list_bond(request):
    # filtering the employee list only which has details
    employee_filter = EmployeeFilter(request.GET, queryset=Employee.objects.filter(details__isnull=False).order_by('-DOJ'))
    
    paginator = Paginator(employee_filter.qs, 10)  
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    return render(request, 'report/employee_Bond_form.html', {'filter': employee_filter, 'page_obj': page_obj})



def BondPaper(request):
    settings_dir = os.path.dirname(__file__)
    project_root = os.path.abspath(os.path.dirname(settings_dir))
    input_file = os.path.join(project_root, 'report/resources/Bond.jrxml')

    if request.method == "POST":
        # Get the selected employee
        employee_list = request.POST.get('employees_list')
        if not employee_list:
            return HttpResponse("No employee selected.", status=400)
        employee = Employee.objects.get(EID=employee_list)
        

        # Create PyReportJasper object
        jasper = PyReportJasper()

        # Define parameters for the report
        param = {
            'EMP_ID': employee.EID
        }

        # Create a temporary file to hold the output
        with tempfile.NamedTemporaryFile(suffix=f'.pdf', delete=False) as temp_file:
            output_file_path = temp_file.name  # Get the temporary file path

        # Configure the report with MySQL database connection
        jasper.config(
            input_file=input_file,
            output_file=output_file_path,  # Save output to the temporary file
            parameters=param,
            output_formats=['pdf'],  # Output in the selected format
            db_connection={
                'driver': 'mysql',
                'username': 'root',
                'password': 'Tonmoy1030',
                'host': 'localhost',
                'database': 'atsdb',
                'port': '3306',
                'jdbc_driver': 'com.mysql.cj.jdbc.Driver',
                'jdbc_url': 'jdbc:mysql://localhost:3306/atsdb'
            }
        )

        # Process the report and generate the output in the temporary file
        jasper.process_report()

        # Read the generated file from the temporary file
        with open(output_file_path, 'rb') as output_file:
            response = HttpResponse(output_file.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{employee.EID}_{employee.name}_Bond.pdf"'
            return response


def EmployeeListReport(request):
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee List"
    
    # Define headers with all necessary fields
    headers = [
        # Employee fields
        'EID', 'Name', 'Designation', 'Department', 'Date of Joining', 'Job Location',
        'Mobile Number', 'Email', 'Business Unit', 'Confirmation Date', 'Active Status',
        # ContactInfo fields
        'Official Mobile', 'Emergency Contact Person', 'Emergency Contact Number',
        'Emergency Person Address', 'Emergency Relation', 'Present Address', 'Permanent Address',
        # PersonalInfo fields
        'Date of Birth', 'Blood Group', 'Father Name', 'Mother Name', 'Religion', 'NID', 'TIN',
        'Marital Status', 'Spouse Name', 'Number of Sons', 'Number of Daughters',
        # Education fields
        'Highest Degree', 'Subject (Highest Degree)', 'Institution (Highest Degree)',
        'Passing Year (Highest Degree)', 'Division/GPA (Highest Degree)', 'Professional Degree',
        'Subject (Professional Degree)', 'Institution (Professional Degree)', 'Passing Year (Professional Degree)',
        # Nominee fields
        'Nominee Name', 'Nominee Father Name', 'Nominee Mother Name', 'Nominee Mobile Number',
        'Relation with Employee', 'Nominee NID', 'Nominee Address'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    form = EmployeeReportForm(request.GET or None)
    
    if form.is_valid():
        # Get filtered data
        unit = form.cleaned_data['unit']
        date_from = form.cleaned_data['date_from']
        date_to = form.cleaned_data['date_to']
        
        
        # Start with base queryset
        employees = Employee.objects.all().select_related('details', 'unit')
        
        # Apply filters
        if unit:
            employees = employees.filter(unit=unit)
        
        date_filter = Q()
        if date_from:
            date_filter &= Q(DOJ__gte=date_from)
        if date_to:
            date_filter &= Q(DOJ__lte=date_to)
        if date_from or date_to:
            employees = employees.filter(date_filter)
        
        # Populate data rows
        for row_idx, employee in enumerate(employees, 2):
            # Employee model fields
            ws.cell(row=row_idx, column=1, value=employee.EID or '')
            ws.cell(row=row_idx, column=2, value=employee.name or '')
            ws.cell(row=row_idx, column=3, value=employee.designation or '')
            ws.cell(row=row_idx, column=4, value=employee.department or '')
            ws.cell(row=row_idx, column=5, value=employee.DOJ.strftime('%Y-%m-%d') if employee.DOJ else '')
            ws.cell(row=row_idx, column=6, value=employee.job_location or '')
            ws.cell(row=row_idx, column=7, value=employee.mobile_no or '')
            ws.cell(row=row_idx, column=8, value=employee.email or '')
            ws.cell(row=row_idx, column=9, value=employee.unit.name if employee.unit else '')
            ws.cell(row=row_idx, column=10, value=employee.confirmation_date.strftime('%Y-%m-%d') if employee.confirmation_date else '')
            ws.cell(row=row_idx, column=11, value='Yes' if employee.active_status else 'No')
            
            # EmployeeDetails fields (ContactInfo, PersonalInfo, Education, Nominee)
            details = getattr(employee, 'details', None)
            
            # ContactInfo fields
            ws.cell(row=row_idx, column=12, value=getattr(details, 'official_mobile', ''))
            ws.cell(row=row_idx, column=13, value=getattr(details, 'emergency_contact_person', ''))
            ws.cell(row=row_idx, column=14, value=getattr(details, 'emergency_contact_no', ''))
            ws.cell(row=row_idx, column=15, value=getattr(details, 'emergency_person_address', ''))
            ws.cell(row=row_idx, column=16, value=getattr(details, 'emer_relation_with_employee', ''))
            
            # Present Address
            present_address = ", ".join(filter(None, [
                getattr(details, 'present_vill', ''),
                getattr(details, 'present_po', ''),
                getattr(details, 'present_ps', ''),
                getattr(details, 'present_dist', '')
            ]))
            ws.cell(row=row_idx, column=17, value=present_address)
            
            # Permanent Address
            permanent_address = ", ".join(filter(None, [
                getattr(details, 'permanent_vill', ''),
                getattr(details, 'permanent_po', ''),
                getattr(details, 'permanent_ps', ''),
                getattr(details, 'permanent_dist', '')
            ]))
            ws.cell(row=row_idx, column=18, value=permanent_address)
            
            # PersonalInfo fields
            ws.cell(row=row_idx, column=19, value=details.date_of_birth.strftime('%Y-%m-%d') if details and details.date_of_birth else '')
            ws.cell(row=row_idx, column=20, value=getattr(details, 'blood_group', ''))
            ws.cell(row=row_idx, column=21, value=getattr(details, 'father_name', ''))
            ws.cell(row=row_idx, column=22, value=getattr(details, 'mother_name', ''))
            ws.cell(row=row_idx, column=23, value=getattr(details, 'religion', ''))
            ws.cell(row=row_idx, column=24, value=getattr(details, 'nid', ''))
            ws.cell(row=row_idx, column=25, value=getattr(details, 'tin', ''))
            ws.cell(row=row_idx, column=26, value=getattr(details, 'marital_status', ''))
            ws.cell(row=row_idx, column=27, value=getattr(details, 'spouse_name', ''))
            ws.cell(row=row_idx, column=28, value=getattr(details, 'no_of_son', ''))
            ws.cell(row=row_idx, column=29, value=getattr(details, 'no_of_daughter', ''))
            
            # Education fields
            ws.cell(row=row_idx, column=30, value=getattr(details, 'highest_degree', ''))
            ws.cell(row=row_idx, column=31, value=getattr(details, 'subject_highest_degree', ''))
            ws.cell(row=row_idx, column=32, value=getattr(details, 'institution_highest_degree', ''))
            ws.cell(row=row_idx, column=33, value=getattr(details, 'passing_year_highest_degree', ''))
            ws.cell(row=row_idx, column=34, value=getattr(details, 'division_or_gpa_highest_degree', ''))
            ws.cell(row=row_idx, column=35, value=getattr(details, 'professional_degree', ''))
            ws.cell(row=row_idx, column=36, value=getattr(details, 'subject_professional_degree', ''))
            ws.cell(row=row_idx, column=37, value=getattr(details, 'institution_professional_degree', ''))
            ws.cell(row=row_idx, column=38, value=getattr(details, 'passing_year_professional_degree', ''))
            
            # Nominee fields
            ws.cell(row=row_idx, column=39, value=getattr(details, 'nominee_name', ''))
            ws.cell(row=row_idx, column=40, value=getattr(details, 'nominee_father_name', ''))
            ws.cell(row=row_idx, column=41, value=getattr(details, 'nominee_mother_name', ''))
            ws.cell(row=row_idx, column=42, value=getattr(details, 'nominee_mobile_no', ''))
            ws.cell(row=row_idx, column=43, value=getattr(details, 'relation_with_employee', ''))
            ws.cell(row=row_idx, column=44, value=getattr(details, 'nominee_nid', ''))
            
            # Nominee Address
            nominee_address = ", ".join(filter(None, [
                getattr(details, 'nominee_vill', ''),
                getattr(details, 'nominee_po', ''),
                getattr(details, 'nominee_ps', ''),
                getattr(details, 'nominee_dist', '')
            ]))
            ws.cell(row=row_idx, column=45, value=nominee_address)
        
        # Adjust column widths dynamically
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Prepare and return response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=Employee_List.xlsx'
        wb.save(response)
        return response
    return render(request, 'report/employee_report_form.html', {'form': form})

def is_admin(user):
    return user.is_authenticated and user.is_staff

@user_passes_test(is_admin, login_url='/accounts/login', redirect_field_name='next')

def EmployeeListReportWithSalary(request):
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee List With Salary"
    
    # Define headers with necessary fields
    headers = [
        'EID', 'Name', 'Designation', 'Department', 
        'Date of Joining', 'Job Location', 'Salary', 'Business Unit'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = openpyxl.styles.Font(bold=True)
    
    form = EmployeeReportForm(request.GET or None)
    
    if form.is_valid():
        # Get filtered data
        unit = form.cleaned_data['unit']
        date_from = form.cleaned_data['date_from']
        date_to = form.cleaned_data['date_to']
        
        # Base queryset with related data
        employees = Employee.objects.all().select_related('unit', 'salary')
        
        # Apply filters
        if unit:
            employees = employees.filter(unit=unit)
        
        date_filter = Q()
        if date_from:
            date_filter &= Q(DOJ__gte=date_from)
        if date_to:
            date_filter &= Q(DOJ__lte=date_to)
        if date_from or date_to:
            employees = employees.filter(date_filter)
        
        # Populate data rows
        for row_idx, employee in enumerate(employees, 2):
            # Safely access salary information
            salary = ''
            if hasattr(employee, 'salary'):
                salary = employee.salary.salary
                
            # Populate cells
            ws.cell(row=row_idx, column=1, value=employee.EID or '')
            ws.cell(row=row_idx, column=2, value=employee.name or '')
            ws.cell(row=row_idx, column=3, value=employee.designation or '')
            ws.cell(row=row_idx, column=4, value=employee.department or '')
            ws.cell(row=row_idx, column=5, value=employee.DOJ.strftime('%Y-%m-%d') if employee.DOJ else '')
            ws.cell(row=row_idx, column=6, value=employee.job_location or '')
            ws.cell(row=row_idx, column=7, value=salary)
            ws.cell(row=row_idx, column=8, value=employee.unit.name if employee.unit else '')
        
        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=Employee_List_Salary.xlsx'
        wb.save(response)
        return response
    
    return render(request, 'report/employee_report_form.html', {'form': form})